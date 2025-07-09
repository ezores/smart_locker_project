"""
Smart Locker System - Flask Backend Application

A comprehensive Flask-based backend for managing equipment borrowing and returning
operations in a smart locker system. This application provides RESTful API endpoints
for the React frontend, handles user authentication, equipment management, and
administrative functions.

Features:
- JWT-based authentication system
- Equipment borrowing and returning operations
- User management (CRUD operations)
- Locker management and status tracking
- Comprehensive activity logging
- Multi-language support (EN, FR, ES, TR)
- Export functionality (Excel, PDF, CSV)
- Real-time statistics and reporting

@author Alp
@date 06/07/2025 - In memory of Yusuf Alpdogan & Mehmet Ugurlu
@version 1.0.0
@description Main Flask application for the Smart Locker Management System
"""

# Standard library imports
import os
import argparse
from datetime import datetime, timedelta
from functools import wraps

# Third-party imports
from flask import Flask, render_template, redirect, url_for, session, request, flash, g, jsonify, Response  # type: ignore[import]
from flask_cors import CORS
from flask_sqlalchemy import SQLAlchemy  # type: ignore[import]
from flask_babel import Babel, gettext, ngettext  # type: ignore[import]
from werkzeug.security import generate_password_hash, check_password_hash  # type: ignore[import]
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user  # type: ignore[import]
import jwt  # type: ignore[import]

# SQLAlchemy imports for advanced queries
from sqlalchemy.orm import aliased  # type: ignore[import]
from sqlalchemy import exists, and_  # type: ignore[import]

# =============================================================================
# APPLICATION CONFIGURATION
# =============================================================================

# Set up base directory and database configuration
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# PostgreSQL Database Configuration
DB_HOST = os.environ.get('DB_HOST', 'localhost')
DB_PORT = os.environ.get('DB_PORT', '5432')
DB_NAME = os.environ.get('DB_NAME', 'smart_locker')
DB_USER = os.environ.get('DB_USER', 'postgres')
DB_PASSWORD = os.environ.get('DB_PASSWORD', 'postgres')

# Initialize Flask application
app = Flask(__name__)

# Enable CORS for cross-origin requests (needed for React frontend)
CORS(app)

# Flask application configuration
app.config['SECRET_KEY'] = 'supersecretkey'  # Change in production
app.config['JWT_SECRET_KEY'] = 'jwt-secret-key-change-in-production'  # Change in production
app.config['SQLALCHEMY_DATABASE_URI'] = f'postgresql://{DB_USER}:{DB_PASSWORD}@{DB_HOST}:{DB_PORT}/{DB_NAME}'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(minutes=30)

# Internationalization (i18n) configuration
app.config['BABEL_DEFAULT_LOCALE'] = 'en'
app.config['BABEL_SUPPORTED_LOCALES'] = ['en', 'fr', 'es', 'tr']
app.config['BABEL_TRANSLATION_DIRECTORIES'] = 'translations'

# Initialize SQLAlchemy database
db = SQLAlchemy(app)

# =============================================================================
# INTERNATIONALIZATION (I18N) SETUP
# =============================================================================

def get_locale():
    """
    Determine the locale for the current request.
    
    Returns:
        str: The locale code (en, fr, es, tr)
    """
    # Check if user has selected a language preference
    if 'language' in session:
        return session['language']
    # Try to guess the language from the user's browser accept header
    return request.accept_languages.best_match(app.config['BABEL_SUPPORTED_LOCALES'])

# Initialize Babel for internationalization
babel = Babel(app, locale_selector=get_locale)

# =============================================================================
# DATABASE MODELS INITIALIZATION
# =============================================================================

def import_models():
    """
    Import and initialize database models.
    
    Returns:
        tuple: Database models and initialization function
    """
    from models import init_models
    return init_models(db)

# Import database models
User, Locker, Item, Log, Borrow, init_db = import_models()

# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def get_duration_since(timestamp):
    """Calculate duration since a timestamp"""
    from datetime import datetime
    now = datetime.now()
    diff = now - timestamp
    
    days = diff.days
    hours = diff.seconds // 3600
    
    if days > 0:
        return f"{days} days"
    elif hours > 0:
        return f"{hours} hours"
    else:
        return "Less than 1 hour"

# =============================================================================
# REQUEST HANDLERS
# =============================================================================

@app.before_request
def before_request():
    """
    Set up global variables before each request.
    """
    g.locale = get_locale()

# --- Language selection ---
@app.route('/language/<language>')
def set_language(language):
    if language in app.config['BABEL_SUPPORTED_LOCALES']:
        session['language'] = language
    return redirect(request.referrer or url_for('main_menu'))

# --- Auth routes ---
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        user = User.query.filter_by(username=username).first()
        if user and check_password_hash(user.password_hash, password):
            session['user_id'] = user.id
            session['username'] = user.username
            session['role'] = user.role
            return redirect(url_for('main_menu'))
        else:
            flash(gettext('Invalid username or password'))
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

# --- Main menu ---
@app.route('/')
def main_menu():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    return render_template('main_menu.html')

# --- Borrow/Return ---
@app.route('/borrow', methods=['GET', 'POST'])
def borrow():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    return render_template('borrow.html')

@app.route('/return', methods=['GET', 'POST'])
def return_item():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    return render_template('return.html')

# --- Admin Dashboard ---
@app.route('/admin')
def admin_dashboard():
    if 'user_id' not in session or session.get('role') != 'admin':
        return redirect(url_for('main_menu'))
    return render_template('admin_dashboard.html')

# --- Users CRUD ---
@app.route('/users')
def users():
    if 'user_id' not in session or session.get('role') != 'admin':
        return redirect(url_for('main_menu'))
    return render_template('users.html')

# --- Items CRUD ---
@app.route('/items')
def items():
    if 'user_id' not in session or session.get('role') != 'admin':
        return redirect(url_for('main_menu'))
    return render_template('items.html')

# --- Lockers CRUD ---
@app.route('/lockers')
def lockers():
    if 'user_id' not in session or session.get('role') != 'admin':
        return redirect(url_for('main_menu'))
    return render_template('lockers.html')

# --- Logs ---
@app.route('/logs')
def logs():
    if 'user_id' not in session or session.get('role') != 'admin':
        return redirect(url_for('main_menu'))
    return render_template('logs.html')

# --- API Routes for React Frontend ---
@app.route('/api/auth/login', methods=['POST'])
def api_login():
    data = request.get_json()
    username = data.get('username')
    password = data.get('password')
    
    user = User.query.filter_by(username=username).first()
    if user and check_password_hash(user.password_hash, password):
        # Create login log
        log = Log(
            user_id=user.id,
            action_type='login'
        )
        db.session.add(log)
        db.session.commit()
        
        # Create JWT token
        token = jwt.encode({
            'user_id': user.id,
            'username': user.username,
            'role': user.role,
            'exp': datetime.utcnow() + timedelta(hours=24)
        }, app.config['JWT_SECRET_KEY'], algorithm='HS256')
        
        return jsonify({
            'token': token,
            'user': {
                'id': user.id,
                'username': user.username,
                'role': user.role
            }
        })
    else:
        return jsonify({'message': 'Invalid username or password'}), 401

@app.route('/api/user/profile')
def api_user_profile():
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'message': 'No token provided'}), 401
    
    token = auth_header.split(' ')[1]
    try:
        payload = jwt.decode(token, app.config['JWT_SECRET_KEY'], algorithms=['HS256'])
        user = User.query.get(payload['user_id'])
        if user:
            return jsonify({
                'id': user.id,
                'username': user.username,
                'role': user.role
            })
        else:
            return jsonify({'message': 'User not found'}), 404
    except jwt.ExpiredSignatureError:
        return jsonify({'message': 'Token expired'}), 401
    except jwt.InvalidTokenError:
        return jsonify({'message': 'Invalid token'}), 401

@app.route('/api/items')
def api_items():
    items = Item.query.all()
    return jsonify([{
        'id': item.id,
        'name': item.name,
        'locker_id': item.locker_id
    } for item in items])

@app.route('/api/lockers')
def api_lockers():
    lockers = Locker.query.all()
    return jsonify([{
        'id': locker.id,
        'number': locker.name,
        'location': f'Location {locker.id}',
        'status': 'available'
    } for locker in lockers])

@app.route('/api/borrow', methods=['POST'])
def api_borrow():
    data = request.get_json()
    rfid_card = data.get('rfid_card')
    item_id = data.get('item_id')
    locker_id = data.get('locker_id')
    
    # Find user by RFID card or use a default user for demo
    user = User.query.filter_by(rfid_tag=rfid_card).first()
    if not user:
        # For demo purposes, use the first student user
        user = User.query.filter_by(role='student').first()
        if not user:
            return jsonify({'error': 'No user found'}), 400
    
    # Check if item is available
    item = Item.query.get(item_id)
    if not item:
        return jsonify({'error': 'Item not found'}), 404
    
    # Check if locker is available
    locker = Locker.query.get(locker_id)
    if not locker or locker.status != 'available':
        return jsonify({'error': 'Locker not available'}), 400
    
    # Create borrow log
    log = Log(
        user_id=user.id,
        item_id=item_id,
        locker_id=locker_id,
        action_type='borrow'
    )
    db.session.add(log)
    
    # Update locker status to occupied
    locker.status = 'occupied'
    
    db.session.commit()
    
    return jsonify({'message': 'Item borrowed successfully'})

@app.route('/api/return', methods=['POST'])
def api_return():
    data = request.get_json()
    rfid_card = data.get('rfid_card')
    item_id = data.get('item_id')
    locker_id = data.get('locker_id')
    
    # Find user by RFID card or use a default user for demo
    user = User.query.filter_by(rfid_tag=rfid_card).first()
    if not user:
        # For demo purposes, use the first student user
        user = User.query.filter_by(role='student').first()
        if not user:
            return jsonify({'error': 'No user found'}), 400
    
    # Check if item exists
    item = Item.query.get(item_id)
    if not item:
        return jsonify({'error': 'Item not found'}), 404
    
    # Check if locker exists
    locker = Locker.query.get(locker_id)
    if not locker:
        return jsonify({'error': 'Locker not found'}), 404
    
    # Create return log
    log = Log(
        user_id=user.id,
        item_id=item_id,
        locker_id=locker_id,
        action_type='return'
    )
    db.session.add(log)
    
    # Update locker status to available
    locker.status = 'available'
    
    db.session.commit()
    
    return jsonify({'message': 'Item returned successfully'})

@app.route('/api/admin/stats')
def api_admin_stats():
    total_users = User.query.count()
    total_items = Item.query.count()
    total_lockers = Locker.query.count()

    # Active borrows: count borrows that have not been returned
    # A borrow is active if there's no return log for the same item and user after the borrow timestamp
    # Get all borrow logs
    borrow_logs = db.session.query(Log).filter(Log.action_type == 'borrow').all()
    
    # Count borrows that don't have a corresponding return
    active_borrows = 0
    for borrow in borrow_logs:
        # Check if there's a return log for this item and user after the borrow timestamp
        has_return = db.session.query(Log).filter(
            and_(
                Log.action_type == 'return',
                Log.item_id == borrow.item_id,
                Log.user_id == borrow.user_id,
                Log.timestamp > borrow.timestamp
            )
        ).first()
        
        if not has_return:
            active_borrows += 1

    return jsonify({
        'totalUsers': total_users,
        'totalItems': total_items,
        'totalLockers': total_lockers,
        'activeBorrows': active_borrows
    })

@app.route('/api/admin/recent-activity')
def api_recent_activity():
    logs = Log.query.order_by(Log.timestamp.desc()).limit(10).all()
    return jsonify([{
        'id': log.id,
        'type': log.action_type,
        'user': User.query.get(log.user_id).username if log.user_id else 'Unknown',
        'item': Item.query.get(log.item_id).name if log.item_id else 'Unknown',
        'timestamp': log.timestamp.isoformat(),
        'status': 'completed'
    } for log in logs])

@app.route('/api/borrowed-items/<identifier>')
def api_borrowed_items(identifier):
    """Get borrowed items for a specific user (by RFID or user ID)"""
    try:
        # Try to find user by RFID card first
        user = User.query.filter_by(rfid_tag=identifier).first()
        
        # If not found by RFID, try by username
        if not user:
            user = User.query.filter_by(username=identifier).first()
        
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        # Get all borrow logs for this user
        borrow_logs = db.session.query(Log).filter(
            and_(
                Log.action_type == 'borrow',
                Log.user_id == user.id
            )
        ).all()
        
        borrowed_items = []
        for borrow_log in borrow_logs:
            # Check if there's a return log for this item after the borrow timestamp
            has_return = db.session.query(Log).filter(
                and_(
                    Log.action_type == 'return',
                    Log.item_id == borrow_log.item_id,
                    Log.user_id == borrow_log.user_id,
                    Log.timestamp > borrow_log.timestamp
                )
            ).first()
            
            # If no return log exists after the borrow, this is still borrowed
            if not has_return:
                item = Item.query.get(borrow_log.item_id)
                locker = Locker.query.get(borrow_log.locker_id) if borrow_log.locker_id else None
                
                borrowed_items.append({
                    'id': borrow_log.id,
                    'item_id': borrow_log.item_id,
                    'name': item.name if item else 'Unknown',
                    'description': item.description if item else '',
                    'locker': {
                        'id': locker.id if locker else None,
                        'name': locker.name if locker else 'Unknown'
                    },
                    'borrowed_at': borrow_log.timestamp.isoformat()
                })
        
        return jsonify(borrowed_items)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/active-borrows')
def api_active_borrows():
    """Get all active borrows (items that have been borrowed but not returned)"""
    try:
        # Get all borrow logs
        borrow_logs = db.session.query(Log).filter(Log.action_type == 'borrow').all()
        
        active_borrows = []
        for borrow_log in borrow_logs:
            # Check if there's a return log for this item and user after the borrow timestamp
            has_return = db.session.query(Log).filter(
                and_(
                    Log.action_type == 'return',
                    Log.item_id == borrow_log.item_id,
                    Log.user_id == borrow_log.user_id,
                    Log.timestamp > borrow_log.timestamp
                )
            ).first()
            
            # If no return log exists after the borrow, this is an active borrow
            if not has_return:
                user = User.query.get(borrow_log.user_id)
                item = Item.query.get(borrow_log.item_id)
                locker = Locker.query.get(borrow_log.locker_id) if borrow_log.locker_id else None
                
                active_borrows.append({
                    'id': borrow_log.id,
                    'user_id': borrow_log.user_id,
                    'user_name': user.username if user else 'Unknown',
                    'item_id': borrow_log.item_id,
                    'item_name': item.name if item else 'Unknown',
                    'locker_id': borrow_log.locker_id,
                    'locker_name': locker.name if locker else 'Unknown',
                    'borrowed_at': borrow_log.timestamp.isoformat(),
                    'description': getattr(borrow_log, 'description', '')
                })
        
        return jsonify(active_borrows)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/reports')
def api_reports():
    """Generate reports based on date range and type"""
    from datetime import datetime, timedelta
    
    report_type = request.args.get('type', 'transactions')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    date_range = request.args.get('range', 'week')
    
    # Parse dates
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d')
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d')
    
    # If no dates provided, use range
    if not start_date or not end_date:
        now = datetime.now()
        end_date = now
        if date_range == 'day':
            start_date = now - timedelta(days=1)
        elif date_range == 'week':
            start_date = now - timedelta(weeks=1)
        elif date_range == 'month':
            start_date = now - timedelta(days=30)
        elif date_range == 'year':
            start_date = now - timedelta(days=365)
        else:
            start_date = now - timedelta(weeks=1)
    
    # Query logs within date range
    logs = Log.query.filter(
        Log.timestamp >= start_date,
        Log.timestamp <= end_date
    ).order_by(Log.timestamp.desc()).all()
    
    # Calculate summary statistics
    total_transactions = len(logs)
    borrows = len([log for log in logs if log.action_type == 'borrow'])
    returns = len([log for log in logs if log.action_type == 'return'])
    unique_users = len(set([log.user_id for log in logs if log.user_id]))
    unique_items = len(set([log.item_id for log in logs if log.item_id]))
    
    # Prepare transaction details
    transactions = []
    for log in logs:
        user = User.query.get(log.user_id)
        item = Item.query.get(log.item_id)
        locker = Locker.query.get(log.locker_id) if log.locker_id else None
        
        transactions.append({
            'id': log.id,
            'user': user.username if user else 'Unknown',
            'item': item.name if item else 'Unknown',
            'action': log.action_type,
            'timestamp': log.timestamp.isoformat(),
            'locker': locker.name if locker else 'Unknown'
        })
    
    return jsonify({
        'summary': {
            'total_transactions': total_transactions,
            'borrows': borrows,
            'returns': returns,
            'unique_users': unique_users,
            'unique_items': unique_items
        },
        'transactions': transactions
    })

@app.route('/api/admin/borrows/export')
def api_export_borrows():
    """Export active borrows in Excel, CSV, or PDF format"""
    from datetime import datetime
    import io
    import csv
    
    format_type = request.args.get('format', 'xlsx')
    
    # Get active borrows using the same logic as the active-borrows endpoint
    borrow_logs = db.session.query(Log).filter(Log.action_type == 'borrow').all()
    
    active_borrows = []
    for borrow_log in borrow_logs:
        # Check if there's a return log for this item and user after the borrow timestamp
        has_return = db.session.query(Log).filter(
            and_(
                Log.action_type == 'return',
                Log.item_id == borrow_log.item_id,
                Log.user_id == borrow_log.user_id,
                Log.timestamp > borrow_log.timestamp
            )
        ).first()
        
        # If no return log exists after the borrow, this is an active borrow
        if not has_return:
            user = User.query.get(borrow_log.user_id)
            item = Item.query.get(borrow_log.item_id)
            locker = Locker.query.get(borrow_log.locker_id) if borrow_log.locker_id else None
            
            active_borrows.append({
                'user': user.username if user else 'Unknown',
                'item': item.name if item else 'Unknown',
                'locker': locker.name if locker else 'Unknown',
                'borrowed_at': borrow_log.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                'duration': get_duration_since(borrow_log.timestamp)
            })
    
    if format_type == 'pdf':
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.lib import colors
            
            # Create PDF
            output = io.BytesIO()
            doc = SimpleDocTemplate(output, pagesize=letter)
            elements = []
            
            # Title
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=30,
                alignment=1  # Center alignment
            )
            title = Paragraph("Active Borrows Report", title_style)
            elements.append(title)
            elements.append(Spacer(1, 20))
            
            # Table data
            table_data = [['User', 'Item', 'Locker', 'Borrowed At', 'Duration']]
            for borrow in active_borrows:
                table_data.append([
                    borrow['user'],
                    borrow['item'],
                    borrow['locker'],
                    borrow['borrowed_at'],
                    borrow['duration']
                ])
            
            # Create table
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            elements.append(table)
            doc.build(elements)
            output.seek(0)
            
            return Response(
                output.getvalue(),
                mimetype='application/pdf',
                headers={"Content-Disposition": f"attachment;filename=active_borrows_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"}
            )
        except ImportError:
            return jsonify({'error': 'PDF generation requires reportlab library'}), 500
    
    elif format_type == 'csv':
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['User', 'Item', 'Locker', 'Borrowed At', 'Duration'])
        
        for borrow in active_borrows:
            writer.writerow([
                borrow['user'],
                borrow['item'],
                borrow['locker'],
                borrow['borrowed_at'],
                borrow['duration']
            ])
        
        output.seek(0)
        return Response(
            output.getvalue(),
            mimetype='text/csv',
            headers={"Content-Disposition": f"attachment;filename=active_borrows_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"}
        )
    
    elif format_type == 'xlsx':
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Active Borrows"
            
            # Headers
            headers = ['User', 'Item', 'Locker', 'Borrowed At', 'Duration']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Data
            for row, borrow in enumerate(active_borrows, 2):
                ws.cell(row=row, column=1, value=borrow['user'])
                ws.cell(row=row, column=2, value=borrow['item'])
                ws.cell(row=row, column=3, value=borrow['locker'])
                ws.cell(row=row, column=4, value=borrow['borrowed_at'])
                ws.cell(row=row, column=5, value=borrow['duration'])
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            return Response(
                output.getvalue(),
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                headers={"Content-Disposition": f"attachment;filename=active_borrows_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"}
            )
        except ImportError:
            return jsonify({'error': 'Excel export requires openpyxl library'}), 500
    
    return jsonify({'error': 'Unsupported format'}), 400

@app.route('/api/admin/export')
def api_export():
    """Export reports in Excel or PDF format"""
    from datetime import datetime, timedelta
    import io
    import csv
    
    report_type = request.args.get('type', 'transactions')
    format_type = request.args.get('format', 'xlsx')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    date_range = request.args.get('range', 'week')
    
    # Parse dates
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d')
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d')
    
    # If no dates provided, use range
    if not start_date or not end_date:
        now = datetime.now()
        end_date = now
        if date_range == 'day':
            start_date = now - timedelta(days=1)
        elif date_range == 'week':
            start_date = now - timedelta(weeks=1)
        elif date_range == 'month':
            start_date = now - timedelta(days=30)
        elif date_range == 'year':
            start_date = now - timedelta(days=365)
        else:
            start_date = now - timedelta(weeks=1)
    
    # Query logs within date range
    logs = Log.query.filter(
        Log.timestamp >= start_date,
        Log.timestamp <= end_date
    ).order_by(Log.timestamp.desc()).all()
    
    if format_type == 'xlsx':
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.worksheet.worksheet import Worksheet
            from openpyxl.utils import get_column_letter
            
            # Create Excel workbook
            wb = Workbook()
            ws = wb.active  # type: ignore
            ws.title = "Smart Locker Report"  # type: ignore
            
            # Style for headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Add headers
            headers = ['User', 'Item', 'Action', 'Timestamp', 'Locker']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)  # type: ignore
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Add data
            for row, log in enumerate(logs, 2):
                user = User.query.get(log.user_id)
                item = Item.query.get(log.item_id)
                locker = Locker.query.get(log.locker_id) if log.locker_id else None
                
                ws.cell(row=row, column=1, value=user.username if user else 'Unknown')  # type: ignore
                ws.cell(row=row, column=2, value=item.name if item else 'Unknown')  # type: ignore
                ws.cell(row=row, column=3, value=log.action_type)  # type: ignore
                ws.cell(row=row, column=4, value=log.timestamp.strftime('%Y-%m-%d %H:%M:%S'))  # type: ignore
                ws.cell(row=row, column=5, value=locker.name if locker else 'Unknown')  # type: ignore
            
            # Auto-adjust column widths
            for col_idx, column in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column), 1):  # type: ignore
                max_length = 0
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width  # type: ignore
            
            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            return Response(
                output.getvalue(),
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                headers={"Content-Disposition": f"attachment;filename=smart_locker_report_{report_type}_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"}
            )
        except ImportError:
            # Fallback to CSV if openpyxl is not available
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(['User', 'Item', 'Action', 'Timestamp', 'Locker'])
            
            for log in logs:
                user = User.query.get(log.user_id)
                item = Item.query.get(log.item_id)
                locker = Locker.query.get(log.locker_id) if log.locker_id else None
                
                writer.writerow([
                    user.username if user else 'Unknown',
                    item.name if item else 'Unknown',
                    log.action_type,
                    log.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                    locker.name if locker else 'Unknown'
                ])
            
            output.seek(0)
            return Response(
                output.getvalue(),
                mimetype='text/csv',
                headers={"Content-Disposition": f"attachment;filename=smart_locker_report_{report_type}_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.csv"}
            )
    
    elif format_type == 'csv':
        # CSV export
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow(['User', 'Item', 'Action', 'Timestamp', 'Locker'])
        
        # Write data
        for log in logs:
            user = User.query.get(log.user_id)
            item = Item.query.get(log.item_id)
            locker = Locker.query.get(log.locker_id) if log.locker_id else None
            
            writer.writerow([
                user.username if user else 'Unknown',
                item.name if item else 'Unknown',
                log.action_type,
                log.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                locker.name if locker else 'Unknown'
            ])
        
        output.seek(0)
        return Response(
            output.getvalue(),
            mimetype='text/csv',
            headers={"Content-Disposition": f"attachment;filename=smart_locker_report_{report_type}_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.csv"}
        )
    
    elif format_type == 'pdf':
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.lib import colors
            
            # Create PDF document
            output = io.BytesIO()
            doc = SimpleDocTemplate(output, pagesize=A4)
            story = []
            
            # Styles
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=30,
                alignment=1  # Center alignment
            )
            
            # Add title
            title = Paragraph(f"Smart Locker System Report<br/>Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}", title_style)
            story.append(title)
            story.append(Spacer(1, 20))
            
            # Prepare table data
            table_data = [['User', 'Item', 'Action', 'Timestamp', 'Locker']]
            
            for log in logs:
                user = User.query.get(log.user_id)
                item = Item.query.get(log.item_id)
                locker = Locker.query.get(log.locker_id) if log.locker_id else None
                
                table_data.append([
                    user.username if user else 'Unknown',
                    item.name if item else 'Unknown',
                    log.action_type,
                    log.timestamp.strftime('%Y-%m-%d %H:%M'),
                    locker.name if locker else 'Unknown'
                ])
            
            # Create table
            table = Table(table_data, colWidths=[1.2*inch, 1.5*inch, 0.8*inch, 1.2*inch, 0.8*inch])
            
            # Style the table
            table_style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ])
            table.setStyle(table_style)
            
            story.append(table)
            
            # Build PDF
            doc.build(story)
            output.seek(0)
            
            return Response(
                output.getvalue(),
                mimetype='application/pdf',
                headers={"Content-Disposition": f"attachment;filename=smart_locker_report_{report_type}_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.pdf"}
            )
        except ImportError:
            # Fallback to text if reportlab is not available
            output = io.StringIO()
            output.write(f"Smart Locker System Report\n")
            output.write(f"Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}\n")
            output.write("=" * 50 + "\n\n")
            output.write("User\tItem\tAction\tTimestamp\tLocker\n")
            output.write("-" * 50 + "\n")
            
            for log in logs:
                user = User.query.get(log.user_id)
                item = Item.query.get(log.item_id)
                locker = Locker.query.get(log.locker_id) if log.locker_id else None
                
                output.write(f"{user.username if user else 'Unknown'}\t")
                output.write(f"{item.name if item else 'Unknown'}\t")
                output.write(f"{log.action_type}\t")
                output.write(f"{log.timestamp.strftime('%Y-%m-%d %H:%M:%S')}\t")
                output.write(f"{locker.name if locker else 'Unknown'}\n")
            
            output.seek(0)
            return Response(
                output.getvalue(),
                mimetype='text/plain',
                headers={"Content-Disposition": f"attachment;filename=smart_locker_report_{report_type}_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.txt"}
            )
    
    return jsonify({'error': 'Unsupported format'}), 400

# --- User Management API Routes ---
@app.route('/api/admin/users')
def api_users():
    """Get all users (admin only)"""
    # Check if user is admin (you might want to add JWT verification here)
    users = User.query.all()
    return jsonify([{
        'id': user.id,
        'username': user.username,
        'role': user.role
    } for user in users])

@app.route('/api/admin/users', methods=['POST'])
def api_create_user():
    """Create a new user (admin only)"""
    data = request.get_json()
    username = data.get('username')
    password = data.get('password')
    role = data.get('role', 'student')
    
    if not username or not password:
        return jsonify({'error': 'Username and password are required'}), 400
    
    # Check if username already exists
    if User.query.filter_by(username=username).first():
        return jsonify({'error': 'Username already exists'}), 400
    
    # Create new user
    user = User(
        username=username,
        password_hash=generate_password_hash(password),
        role=role
    )
    db.session.add(user)
    db.session.commit()
    
    return jsonify({
        'id': user.id,
        'username': user.username,
        'role': user.role
    }), 201

@app.route('/api/admin/users/<int:user_id>', methods=['PUT'])
def api_update_user(user_id):
    """Update a user (admin only)"""
    data = request.get_json()
    username = data.get('username')
    password = data.get('password')
    role = data.get('role')
    
    user = User.query.get(user_id)
    if not user:
        return jsonify({'error': 'User not found'}), 404
    
    if username:
        # Check if username already exists (excluding current user)
        existing_user = User.query.filter_by(username=username).first()
        if existing_user and existing_user.id != user_id:
            return jsonify({'error': 'Username already exists'}), 400
        user.username = username
    
    if password:
        user.password_hash = generate_password_hash(password)
    
    if role:
        user.role = role
    
    db.session.commit()
    
    return jsonify({
        'id': user.id,
        'username': user.username,
        'role': user.role
    })

@app.route('/api/admin/users/<int:user_id>', methods=['DELETE'])
def api_delete_user(user_id):
    """Delete a user (admin only)"""
    user = User.query.get(user_id)
    if not user:
        return jsonify({'error': 'User not found'}), 404
    
    # Prevent deleting admin users
    if user.role == 'admin':
        return jsonify({'error': 'Cannot delete admin users'}), 400
    
    db.session.delete(user)
    db.session.commit()
    
    return jsonify({'message': 'User deleted successfully'})

# Items Management API
@app.route('/api/admin/items', methods=['POST'])
def api_create_item():
    """Create a new item (admin only)"""
    data = request.get_json()
    name = data.get('name')
    locker_id = data.get('locker_id')
    description = data.get('description', '')
    
    if not name:
        return jsonify({'error': 'Item name is required'}), 400
    
    # Create new item
    item = Item(
        name=name,
        locker_id=locker_id if locker_id else None,
        description=description
    )
    db.session.add(item)
    db.session.commit()
    
    return jsonify({
        'id': item.id,
        'name': item.name,
        'locker_id': item.locker_id,
        'description': item.description
    }), 201

@app.route('/api/admin/items/<int:item_id>', methods=['PUT'])
def api_update_item(item_id):
    """Update an item (admin only)"""
    item = Item.query.get_or_404(item_id)
    data = request.get_json()
    
    if 'name' in data:
        item.name = data['name']
    if 'locker_id' in data:
        item.locker_id = data['locker_id'] if data['locker_id'] else None
    if 'description' in data:
        item.description = data['description']
    
    db.session.commit()
    
    return jsonify({
        'id': item.id,
        'name': item.name,
        'locker_id': item.locker_id,
        'description': item.description
    })

@app.route('/api/admin/items/<int:item_id>', methods=['DELETE'])
def api_delete_item(item_id):
    """Delete an item (admin only)"""
    item = Item.query.get_or_404(item_id)
    db.session.delete(item)
    db.session.commit()
    
    return jsonify({'message': 'Item deleted successfully'})

# Lockers Management API
@app.route('/api/admin/lockers', methods=['POST'])
def api_create_locker():
    """Create a new locker (admin only)"""
    data = request.get_json()
    name = data.get('name')
    location = data.get('location', '')
    status = data.get('status', 'available')
    
    if not name:
        return jsonify({'error': 'Locker name is required'}), 400
    
    # Create new locker
    locker = Locker(
        name=name,
        location=location,
        status=status
    )
    db.session.add(locker)
    db.session.commit()
    
    return jsonify({
        'id': locker.id,
        'name': locker.name,
        'location': locker.location,
        'status': locker.status
    }), 201

@app.route('/api/admin/lockers/<int:locker_id>', methods=['PUT'])
def api_update_locker(locker_id):
    """Update a locker (admin only)"""
    locker = Locker.query.get_or_404(locker_id)
    data = request.get_json()
    
    if 'name' in data:
        locker.name = data['name']
    if 'location' in data:
        locker.location = data['location']
    if 'status' in data:
        locker.status = data['status']
    
    db.session.commit()
    
    return jsonify({
        'id': locker.id,
        'name': locker.name,
        'location': locker.location,
        'status': locker.status
    })

@app.route('/api/admin/lockers/<int:locker_id>', methods=['DELETE'])
def api_delete_locker(locker_id):
    """Delete a locker (admin only)"""
    locker = Locker.query.get_or_404(locker_id)
    db.session.delete(locker)
    db.session.commit()
    
    return jsonify({'message': 'Locker deleted successfully'})

# Logs Export API
@app.route('/api/admin/logs/export')
def api_export_logs():
    """Export logs in various formats (admin only)"""
    format_type = request.args.get('format', 'csv')
    
    # Get all logs with user, item, and locker information
    logs = db.session.query(Log, User.username, Item.name.label('item_name'), Locker.name.label('locker_name'))\
        .outerjoin(User, Log.user_id == User.id)\
        .outerjoin(Item, Log.item_id == Item.id)\
        .outerjoin(Locker, Log.locker_id == Locker.id)\
        .order_by(Log.timestamp.desc()).all()
    
    if format_type == 'csv':
        import csv
        from io import StringIO
        
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(['ID', 'Action', 'User', 'Item', 'Locker', 'Description', 'Timestamp'])
        
        for log, username, item_name, locker_name in logs:
            writer.writerow([
                log.id,
                log.action_type,
                username or 'Unknown',
                item_name or 'N/A',
                locker_name or 'N/A',
                '',  # Log model doesn't have description field
                log.timestamp.strftime('%Y-%m-%d %H:%M:%S')
            ])
        
        output.seek(0)
        return Response(
            output.getvalue(),
            mimetype='text/csv',
            headers={'Content-Disposition': f'attachment; filename=logs_{datetime.now().strftime("%Y%m%d")}.csv'}
        )
    
    elif format_type == 'excel':
        try:
            import pandas as pd
            from io import BytesIO
            
            # Prepare data for pandas
            data = []
            for log, username, item_name, locker_name in logs:
                data.append({
                    'ID': log.id,
                    'Action': log.action_type,
                    'User': username or 'Unknown',
                    'Item': item_name or 'N/A',
                    'Locker': locker_name or 'N/A',
                    'Description': '',  # Log model doesn't have description field
                    'Timestamp': log.timestamp.strftime('%Y-%m-%d %H:%M:%S')
                })
            
            df = pd.DataFrame(data)
            
            # Create Excel file in memory
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Logs', index=False)
            
            output.seek(0)
            return Response(
                output.getvalue(),
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                headers={'Content-Disposition': f'attachment; filename=logs_{datetime.now().strftime("%Y%m%d")}.xlsx'}
            )
        except ImportError:
            return jsonify({'error': 'pandas and openpyxl are required for Excel export'}), 500
    
    elif format_type == 'pdf':
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.lib import colors
            
            # Create PDF
            output = io.BytesIO()
            doc = SimpleDocTemplate(output, pagesize=letter)
            elements = []
            
            # Title
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=30,
                alignment=1  # Center alignment
            )
            title = Paragraph("System Logs Report", title_style)
            elements.append(title)
            elements.append(Spacer(1, 20))
            
            # Table data
            table_data = [['ID', 'Action', 'User', 'Item', 'Locker', 'Timestamp']]
            for log, username, item_name, locker_name in logs:
                table_data.append([
                    str(log.id),
                    log.action_type,
                    username or 'Unknown',
                    item_name or 'N/A',
                    locker_name or 'N/A',
                    log.timestamp.strftime('%Y-%m-%d %H:%M:%S')
                ])
            
            # Create table
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.beige, colors.white])
            ]))
            
            elements.append(table)
            doc.build(elements)
            output.seek(0)
            
            return Response(
                output.getvalue(),
                mimetype='application/pdf',
                headers={"Content-Disposition": f"attachment;filename=logs_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"}
            )
        except ImportError:
            return jsonify({'error': 'PDF generation requires reportlab library'}), 500
    
    else:
        return jsonify({'error': 'Unsupported format'}), 400

@app.route('/api/health')
def api_health():
    """Health check endpoint"""
    try:
        # Test database connection
        db.session.execute('SELECT 1')
        db_status = 'healthy'
    except Exception as e:
        db_status = f'unhealthy: {str(e)}'
    
    return jsonify({
        'status': 'healthy',
        'timestamp': datetime.now().isoformat(),
        'database': db_status,
        'version': '1.0.0'
    })

@app.route('/api/logs')
def api_logs():
    """Get all logs"""
    logs = Log.query.order_by(Log.timestamp.desc()).all()
    
    logs_data = []
    for log in logs:
        user = User.query.get(log.user_id)
        item = Item.query.get(log.item_id)
        locker = Locker.query.get(log.locker_id) if log.locker_id else None
        
        logs_data.append({
            'id': log.id,
            'action_type': log.action_type,
            'user_id': log.user_id,
            'user_name': user.username if user else 'Unknown',
            'item_id': log.item_id,
            'item_name': item.name if item else 'Unknown',
            'locker_id': log.locker_id,
            'locker_name': locker.name if locker else 'Unknown',
            'timestamp': log.timestamp.isoformat(),
            'description': getattr(log, 'description', '')
        })
    
    return jsonify(logs_data)

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Smart Locker System')
    parser.add_argument('--port', type=int, default=5050, help='Port to run the server on (default: 5050)')
    parser.add_argument('--host', type=str, default='0.0.0.0', help='Host to bind to (default: 0.0.0.0)')
    parser.add_argument('--demo', action='store_true', help='Load comprehensive demo data for testing')
    parser.add_argument('--simple-demo', action='store_true', help='Load minimal demo data for testing')
    parser.add_argument('--reset-db', action='store_true', help='Reset database and load demo data')
    parser.add_argument('--verbose', action='store_true', help='Enable verbose logging')
    args = parser.parse_args()
    
    with app.app_context():
        init_db()
        
        # Load demo data if requested
        if args.demo or args.simple_demo or args.reset_db:
            try:
                if args.reset_db:
                    print("Resetting database and loading demo data...")
                    # Clear all data
                    db.session.query(Borrow).delete()
                    db.session.query(Log).delete()
                    db.session.query(Item).delete()
                    db.session.query(Locker).delete()
                    db.session.query(User).delete()
                    db.session.commit()
                
                if args.simple_demo:
                    print("Loading simple demo data...")
                    # Create minimal demo data
                    admin = User(
                        username='admin',
                        password_hash=generate_password_hash('admin123'),
                        role='admin',
                        email='admin@ets.com',
                        first_name='Admin',
                        last_name='User'
                    )
                    student = User(
                        username='student',
                        password_hash=generate_password_hash('student123'),
                        role='student',
                        email='student@ets.com',
                        first_name='Student',
                        last_name='User',
                        student_id='2024001'
                    )
                    db.session.add(admin)
                    db.session.add(student)
                    
                    locker1 = Locker(name='Locker A1', location='Building A', capacity=10)
                    locker2 = Locker(name='Locker B1', location='Building B', capacity=8)
                    db.session.add(locker1)
                    db.session.add(locker2)
                    
                    item1 = Item(name='Laptop', description='MacBook Pro', category='electronics', condition='excellent', serial_number='MBP001', locker=locker1)
                    item2 = Item(name='Tablet', description='iPad Pro', category='electronics', condition='good', serial_number='IPP001', locker=locker2)
                    db.session.add(item1)
                    db.session.add(item2)
                    
                    db.session.commit()
                    print("Simple demo data loaded successfully!")
                else:
                    print("Loading comprehensive demo data...")
                    # Use the comprehensive demo data from models.py
                    from models import generate_dummy_data
                    generate_dummy_data()
                    print("Comprehensive demo data loaded successfully!")
            except Exception as e:
                print(f"Error loading demo data: {e}")
    
    print(f"Starting Smart Locker System on http://{args.host}:{args.port}")
    app.run(debug=True, host=args.host, port=args.port) 