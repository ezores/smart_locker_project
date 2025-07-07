"""
Smart Locker System - Flask Backend Application

@author Alp
@date 2024-12-XX
@description Main Flask application for the Smart Locker Management System
"""

import os
import argparse
from flask import Flask, render_template, redirect, url_for, session, request, flash, g, jsonify, Response  # type: ignore[import]
from flask_cors import CORS
from flask_sqlalchemy import SQLAlchemy  # type: ignore[import]
from flask_babel import Babel, gettext, ngettext  # type: ignore[import]
from werkzeug.security import generate_password_hash, check_password_hash  # type: ignore[import]
import jwt  # type: ignore[import]
from datetime import datetime, timedelta
from datetime import timedelta

# Set up base directory
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_DIR = os.path.join(BASE_DIR, 'db')
DB_PATH = os.path.join(DB_DIR, 'locker.db')

if not os.path.exists(DB_DIR):
    os.makedirs(DB_DIR)

app = Flask(__name__)
CORS(app)  # Enable CORS for all routes
app.config['SECRET_KEY'] = 'supersecretkey'
app.config['JWT_SECRET_KEY'] = 'jwt-secret-key-change-in-production'
app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{DB_PATH}'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(minutes=30)

# Babel configuration
app.config['BABEL_DEFAULT_LOCALE'] = 'en'
app.config['BABEL_SUPPORTED_LOCALES'] = ['en', 'fr', 'es', 'tr']
app.config['BABEL_TRANSLATION_DIRECTORIES'] = 'translations'

db = SQLAlchemy(app)

def get_locale():
    # Check if user has selected a language
    if 'language' in session:
        return session['language']
    # Try to guess the language from the user accept header
    return request.accept_languages.best_match(app.config['BABEL_SUPPORTED_LOCALES'])

babel = Babel(app, locale_selector=get_locale)

def import_models():
    from models import init_models
    return init_models(db)

User, Locker, Item, Log, init_db = import_models()

@app.before_request
def before_request():
    g.locale = get_locale()

# --- Language selection ---
@app.route('/language/<language>')
def set_language(language):
    if language in app.config['BABEL_SUPPORTED_LOCALES']:
        session['language'] = language
    return redirect(request.referrer or url_for('main_menu'))

# --- Auth routes ---
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        user = User.query.filter_by(username=username).first()
        if user and check_password_hash(user.password_hash, password):
            session['user_id'] = user.id
            session['username'] = user.username
            session['role'] = user.role
            return redirect(url_for('main_menu'))
        else:
            flash(gettext('Invalid username or password'))
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

# --- Main menu ---
@app.route('/')
def main_menu():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    return render_template('main_menu.html')

# --- Borrow/Return ---
@app.route('/borrow', methods=['GET', 'POST'])
def borrow():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    return render_template('borrow.html')

@app.route('/return', methods=['GET', 'POST'])
def return_item():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    return render_template('return.html')

# --- Admin Dashboard ---
@app.route('/admin')
def admin_dashboard():
    if 'user_id' not in session or session.get('role') != 'admin':
        return redirect(url_for('main_menu'))
    return render_template('admin_dashboard.html')

# --- Users CRUD ---
@app.route('/users')
def users():
    if 'user_id' not in session or session.get('role') != 'admin':
        return redirect(url_for('main_menu'))
    return render_template('users.html')

# --- Items CRUD ---
@app.route('/items')
def items():
    if 'user_id' not in session or session.get('role') != 'admin':
        return redirect(url_for('main_menu'))
    return render_template('items.html')

# --- Lockers CRUD ---
@app.route('/lockers')
def lockers():
    if 'user_id' not in session or session.get('role') != 'admin':
        return redirect(url_for('main_menu'))
    return render_template('lockers.html')

# --- Logs ---
@app.route('/logs')
def logs():
    if 'user_id' not in session or session.get('role') != 'admin':
        return redirect(url_for('main_menu'))
    return render_template('logs.html')

# --- API Routes for React Frontend ---
@app.route('/api/auth/login', methods=['POST'])
def api_login():
    data = request.get_json()
    username = data.get('username')
    password = data.get('password')
    
    user = User.query.filter_by(username=username).first()
    if user and check_password_hash(user.password_hash, password):
        # Create JWT token
        token = jwt.encode({
            'user_id': user.id,
            'username': user.username,
            'role': user.role,
            'exp': datetime.utcnow() + timedelta(hours=24)
        }, app.config['JWT_SECRET_KEY'], algorithm='HS256')
        
        return jsonify({
            'token': token,
            'user': {
                'id': user.id,
                'username': user.username,
                'role': user.role
            }
        })
    else:
        return jsonify({'message': 'Invalid username or password'}), 401

@app.route('/api/user/profile')
def api_user_profile():
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'message': 'No token provided'}), 401
    
    token = auth_header.split(' ')[1]
    try:
        payload = jwt.decode(token, app.config['JWT_SECRET_KEY'], algorithms=['HS256'])
        user = User.query.get(payload['user_id'])
        if user:
            return jsonify({
                'id': user.id,
                'username': user.username,
                'role': user.role
            })
        else:
            return jsonify({'message': 'User not found'}), 404
    except jwt.ExpiredSignatureError:
        return jsonify({'message': 'Token expired'}), 401
    except jwt.InvalidTokenError:
        return jsonify({'message': 'Invalid token'}), 401

@app.route('/api/items')
def api_items():
    items = Item.query.all()
    return jsonify([{
        'id': item.id,
        'name': item.name,
        'locker_id': item.locker_id
    } for item in items])

@app.route('/api/lockers')
def api_lockers():
    lockers = Locker.query.all()
    return jsonify([{
        'id': locker.id,
        'number': locker.name,
        'location': f'Location {locker.id}',
        'status': 'available'
    } for locker in lockers])

@app.route('/api/borrow', methods=['POST'])
def api_borrow():
    data = request.get_json()
    # Mock borrow functionality
    return jsonify({'message': 'Item borrowed successfully'})

@app.route('/api/return', methods=['POST'])
def api_return():
    data = request.get_json()
    # Mock return functionality
    return jsonify({'message': 'Item returned successfully'})

@app.route('/api/admin/stats')
def api_admin_stats():
    return jsonify({
        'totalUsers': User.query.count(),
        'totalItems': Item.query.count(),
        'totalLockers': Locker.query.count(),
        'activeBorrows': Log.query.filter_by(action_type='borrow').count()
    })

@app.route('/api/admin/recent-activity')
def api_recent_activity():
    logs = Log.query.order_by(Log.timestamp.desc()).limit(10).all()
    return jsonify([{
        'id': log.id,
        'type': log.action_type,
        'user': User.query.get(log.user_id).username if log.user_id else 'Unknown',
        'item': Item.query.get(log.item_id).name if log.item_id else 'Unknown',
        'timestamp': log.timestamp.isoformat(),
        'status': 'completed'
    } for log in logs])

@app.route('/api/admin/reports')
def api_reports():
    """Generate reports based on date range and type"""
    from datetime import datetime, timedelta
    
    report_type = request.args.get('type', 'transactions')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    date_range = request.args.get('range', 'week')
    
    # Parse dates
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d')
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d')
    
    # If no dates provided, use range
    if not start_date or not end_date:
        now = datetime.now()
        end_date = now
        if date_range == 'day':
            start_date = now - timedelta(days=1)
        elif date_range == 'week':
            start_date = now - timedelta(weeks=1)
        elif date_range == 'month':
            start_date = now - timedelta(days=30)
        elif date_range == 'year':
            start_date = now - timedelta(days=365)
        else:
            start_date = now - timedelta(weeks=1)
    
    # Query logs within date range
    logs = Log.query.filter(
        Log.timestamp >= start_date,
        Log.timestamp <= end_date
    ).order_by(Log.timestamp.desc()).all()
    
    # Calculate summary statistics
    total_transactions = len(logs)
    borrows = len([log for log in logs if log.action_type == 'borrow'])
    returns = len([log for log in logs if log.action_type == 'return'])
    unique_users = len(set([log.user_id for log in logs if log.user_id]))
    unique_items = len(set([log.item_id for log in logs if log.item_id]))
    
    # Prepare transaction details
    transactions = []
    for log in logs:
        user = User.query.get(log.user_id)
        item = Item.query.get(log.item_id)
        locker = Locker.query.get(log.locker_id) if log.locker_id else None
        
        transactions.append({
            'id': log.id,
            'user': user.username if user else 'Unknown',
            'item': item.name if item else 'Unknown',
            'action': log.action_type,
            'timestamp': log.timestamp.isoformat(),
            'locker': locker.name if locker else 'Unknown'
        })
    
    return jsonify({
        'summary': {
            'total_transactions': total_transactions,
            'borrows': borrows,
            'returns': returns,
            'unique_users': unique_users,
            'unique_items': unique_items
        },
        'transactions': transactions
    })

@app.route('/api/admin/export')
def api_export():
    """Export reports in Excel or PDF format"""
    from datetime import datetime, timedelta
    import io
    import csv
    
    report_type = request.args.get('type', 'transactions')
    format_type = request.args.get('format', 'xlsx')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    date_range = request.args.get('range', 'week')
    
    # Parse dates
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d')
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d')
    
    # If no dates provided, use range
    if not start_date or not end_date:
        now = datetime.now()
        end_date = now
        if date_range == 'day':
            start_date = now - timedelta(days=1)
        elif date_range == 'week':
            start_date = now - timedelta(weeks=1)
        elif date_range == 'month':
            start_date = now - timedelta(days=30)
        elif date_range == 'year':
            start_date = now - timedelta(days=365)
        else:
            start_date = now - timedelta(weeks=1)
    
    # Query logs within date range
    logs = Log.query.filter(
        Log.timestamp >= start_date,
        Log.timestamp <= end_date
    ).order_by(Log.timestamp.desc()).all()
    
    if format_type == 'xlsx':
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.worksheet.worksheet import Worksheet
            from openpyxl.utils import get_column_letter
            
            # Create Excel workbook
            wb = Workbook()
            ws = wb.active  # type: ignore
            ws.title = "Smart Locker Report"  # type: ignore
            
            # Style for headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Add headers
            headers = ['User', 'Item', 'Action', 'Timestamp', 'Locker']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)  # type: ignore
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Add data
            for row, log in enumerate(logs, 2):
                user = User.query.get(log.user_id)
                item = Item.query.get(log.item_id)
                locker = Locker.query.get(log.locker_id) if log.locker_id else None
                
                ws.cell(row=row, column=1, value=user.username if user else 'Unknown')  # type: ignore
                ws.cell(row=row, column=2, value=item.name if item else 'Unknown')  # type: ignore
                ws.cell(row=row, column=3, value=log.action_type)  # type: ignore
                ws.cell(row=row, column=4, value=log.timestamp.strftime('%Y-%m-%d %H:%M:%S'))  # type: ignore
                ws.cell(row=row, column=5, value=locker.name if locker else 'Unknown')  # type: ignore
            
            # Auto-adjust column widths
            for col_idx, column in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column), 1):  # type: ignore
                max_length = 0
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width  # type: ignore
            
            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            return Response(
                output.getvalue(),
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                headers={"Content-Disposition": f"attachment;filename=smart_locker_report_{report_type}_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"}
            )
        except ImportError:
            # Fallback to CSV if openpyxl is not available
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(['User', 'Item', 'Action', 'Timestamp', 'Locker'])
            
            for log in logs:
                user = User.query.get(log.user_id)
                item = Item.query.get(log.item_id)
                locker = Locker.query.get(log.locker_id) if log.locker_id else None
                
                writer.writerow([
                    user.username if user else 'Unknown',
                    item.name if item else 'Unknown',
                    log.action_type,
                    log.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                    locker.name if locker else 'Unknown'
                ])
            
            output.seek(0)
            return Response(
                output.getvalue(),
                mimetype='text/csv',
                headers={"Content-Disposition": f"attachment;filename=smart_locker_report_{report_type}_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.csv"}
            )
    
    elif format_type == 'csv':
        # CSV export
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow(['User', 'Item', 'Action', 'Timestamp', 'Locker'])
        
        # Write data
        for log in logs:
            user = User.query.get(log.user_id)
            item = Item.query.get(log.item_id)
            locker = Locker.query.get(log.locker_id) if log.locker_id else None
            
            writer.writerow([
                user.username if user else 'Unknown',
                item.name if item else 'Unknown',
                log.action_type,
                log.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                locker.name if locker else 'Unknown'
            ])
        
        output.seek(0)
        return Response(
            output.getvalue(),
            mimetype='text/csv',
            headers={"Content-Disposition": f"attachment;filename=smart_locker_report_{report_type}_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.csv"}
        )
    
    elif format_type == 'pdf':
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.lib import colors
            
            # Create PDF document
            output = io.BytesIO()
            doc = SimpleDocTemplate(output, pagesize=A4)
            story = []
            
            # Styles
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=30,
                alignment=1  # Center alignment
            )
            
            # Add title
            title = Paragraph(f"Smart Locker System Report<br/>Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}", title_style)
            story.append(title)
            story.append(Spacer(1, 20))
            
            # Prepare table data
            table_data = [['User', 'Item', 'Action', 'Timestamp', 'Locker']]
            
            for log in logs:
                user = User.query.get(log.user_id)
                item = Item.query.get(log.item_id)
                locker = Locker.query.get(log.locker_id) if log.locker_id else None
                
                table_data.append([
                    user.username if user else 'Unknown',
                    item.name if item else 'Unknown',
                    log.action_type,
                    log.timestamp.strftime('%Y-%m-%d %H:%M'),
                    locker.name if locker else 'Unknown'
                ])
            
            # Create table
            table = Table(table_data, colWidths=[1.2*inch, 1.5*inch, 0.8*inch, 1.2*inch, 0.8*inch])
            
            # Style the table
            table_style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ])
            table.setStyle(table_style)
            
            story.append(table)
            
            # Build PDF
            doc.build(story)
            output.seek(0)
            
            return Response(
                output.getvalue(),
                mimetype='application/pdf',
                headers={"Content-Disposition": f"attachment;filename=smart_locker_report_{report_type}_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.pdf"}
            )
        except ImportError:
            # Fallback to text if reportlab is not available
            output = io.StringIO()
            output.write(f"Smart Locker System Report\n")
            output.write(f"Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}\n")
            output.write("=" * 50 + "\n\n")
            output.write("User\tItem\tAction\tTimestamp\tLocker\n")
            output.write("-" * 50 + "\n")
            
            for log in logs:
                user = User.query.get(log.user_id)
                item = Item.query.get(log.item_id)
                locker = Locker.query.get(log.locker_id) if log.locker_id else None
                
                output.write(f"{user.username if user else 'Unknown'}\t")
                output.write(f"{item.name if item else 'Unknown'}\t")
                output.write(f"{log.action_type}\t")
                output.write(f"{log.timestamp.strftime('%Y-%m-%d %H:%M:%S')}\t")
                output.write(f"{locker.name if locker else 'Unknown'}\n")
            
            output.seek(0)
            return Response(
                output.getvalue(),
                mimetype='text/plain',
                headers={"Content-Disposition": f"attachment;filename=smart_locker_report_{report_type}_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.txt"}
            )
    
    return jsonify({'error': 'Unsupported format'}), 400

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Smart Locker System')
    parser.add_argument('--port', type=int, default=5050, help='Port to run the server on (default: 5050)')
    parser.add_argument('--host', type=str, default='0.0.0.0', help='Host to bind to (default: 0.0.0.0)')
    parser.add_argument('--demo', action='store_true', help='Load demo data for testing')
    args = parser.parse_args()
    
    with app.app_context():
        init_db()
        
        # Load demo data if requested
        if args.demo:
            try:
                from demo_data import create_demo_data
                create_demo_data(db, User, Locker, Item, Log)
                print("✅ Demo data loaded successfully!")
            except Exception as e:
                print(f"❌ Error loading demo data: {e}")
    
    print(f"Starting Smart Locker System on http://{args.host}:{args.port}")
    app.run(debug=True, host=args.host, port=args.port) 