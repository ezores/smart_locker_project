import io
import csv
import json
import logging
from datetime import datetime, timedelta
from typing import List, Dict, Any, Optional
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

logger = logging.getLogger(__name__)

class ExportManager:
    def __init__(self):
        self.styles = getSampleStyleSheet()
        self.setup_custom_styles()
    
    def setup_custom_styles(self):
        """Setup custom paragraph styles for PDF reports"""
        self.title_style = ParagraphStyle(
            'CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1,  # Center alignment
            textColor=colors.darkblue
        )
        
        self.heading_style = ParagraphStyle(
            'CustomHeading',
            parent=self.styles['Heading2'],
            fontSize=12,
            spaceAfter=12,
            textColor=colors.darkblue
        )
        
        self.normal_style = ParagraphStyle(
            'CustomNormal',
            parent=self.styles['Normal'],
            fontSize=10,
            spaceAfter=6
        )
    
    def export_csv(self, data: List[Dict], filename: str = None) -> str:
        """Export data to CSV format"""
        if not data:
            return ""
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write header
        if data:
            writer.writerow(data[0].keys())
        
        # Write data
        for row in data:
            writer.writerow(row.values())
        
        csv_content = output.getvalue()
        output.close()
        
        if filename:
            with open(filename, 'w', newline='', encoding='utf-8') as f:
                f.write(csv_content)
            logger.info(f"CSV exported to {filename}")
        
        return csv_content
    
    def export_excel(self, data: List[Dict], filename: str = None, sheet_name: str = "Data") -> bytes:
        """Export data to Excel format"""
        if not data:
            return b""
        
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Write header
        if data:
            headers = list(data[0].keys())
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Write data
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data.values(), 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        excel_content = output.getvalue()
        output.close()
        
        if filename:
            with open(filename, 'wb') as f:
                f.write(excel_content)
            logger.info(f"Excel exported to {filename}")
        
        return excel_content
    
    def export_pdf(self, title: str, sections: List[Dict], filename: str = None) -> bytes:
        """Export data to PDF format with sections"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        story = []
        
        # Title
        story.append(Paragraph(title, self.title_style))
        story.append(Spacer(1, 12))
        
        # Add timestamp
        timestamp = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        story.append(Paragraph(timestamp, self.normal_style))
        story.append(Spacer(1, 20))
        
        # Process sections
        for section in sections:
            # Section title
            if 'title' in section:
                story.append(Paragraph(section['title'], self.heading_style))
                story.append(Spacer(1, 12))
            
            # Section content
            if 'content' in section:
                if isinstance(section['content'], str):
                    story.append(Paragraph(section['content'], self.normal_style))
                elif isinstance(section['content'], list):
                    # Create table from list of dictionaries
                    if section['content'] and isinstance(section['content'][0], dict):
                        headers = list(section['content'][0].keys())
                        table_data = [headers]
                        for row in section['content']:
                            table_data.append([str(row.get(header, '')) for header in headers])
                        
                        table = Table(table_data)
                        table.setStyle(TableStyle([
                            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                            ('FONTSIZE', (0, 0), (-1, 0), 10),
                            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                            ('GRID', (0, 0), (-1, -1), 1, colors.black)
                        ]))
                        story.append(table)
                    else:
                        # Simple list
                        for item in section['content']:
                            story.append(Paragraph(f"• {item}", self.normal_style))
            
            story.append(Spacer(1, 20))
        
        # Build PDF
        doc.build(story)
        pdf_content = buffer.getvalue()
        buffer.close()
        
        if filename:
            with open(filename, 'wb') as f:
                f.write(pdf_content)
            logger.info(f"PDF exported to {filename}")
        
        return pdf_content
    
    def create_system_report(self, db_session, report_type: str = "comprehensive") -> Dict[str, Any]:
        """Create comprehensive system report"""
        from app import User, Locker, Item, Borrow, Log, Return
        
        report_data = {
            "timestamp": datetime.now(),
            "report_type": report_type,
            "summary": {},
            "details": {}
        }
        
        # Get basic statistics
        total_users = User.query.count()
        total_lockers = Locker.query.count()
        total_items = Item.query.count()
        active_borrows = Borrow.query.filter_by(status='active').count()
        total_logs = Log.query.count()
        
        report_data["summary"] = {
            "total_users": total_users,
            "total_lockers": total_lockers,
            "total_items": total_items,
            "active_borrows": active_borrows,
            "total_logs": total_logs
        }
        
        if report_type == "comprehensive":
            # Get detailed data
            users = [user.to_dict() for user in User.query.all()]
            lockers = [locker.to_dict() for locker in Locker.query.all()]
            items = [item.to_dict() for item in Item.query.all()]
            borrows = [borrow.to_dict() for borrow in Borrow.query.all()]
            recent_logs = [log.to_dict() for log in Log.query.order_by(Log.timestamp.desc()).limit(100).all()]
            
            report_data["details"] = {
                "users": users,
                "lockers": lockers,
                "items": items,
                "borrows": borrows,
                "recent_logs": recent_logs
            }
        
        return report_data
    
    def export_system_report(self, db_session, format_type: str = "pdf", report_type: str = "comprehensive") -> bytes:
        """Export system report in specified format"""
        report_data = self.create_system_report(db_session, report_type)
        
        if format_type == "csv":
            # Export as CSV
            if report_type == "comprehensive":
                # Export multiple CSV files
                csv_data = {}
                for key, data in report_data["details"].items():
                    csv_data[f"{key}.csv"] = self.export_csv(data)
                return csv_data
            else:
                # Export summary as CSV
                summary_data = [{"metric": k, "value": v} for k, v in report_data["summary"].items()]
                return self.export_csv(summary_data)
        
        elif format_type == "excel":
            # Export as Excel
            if report_type == "comprehensive":
                # Create multi-sheet Excel file
                wb = Workbook()
                
                # Summary sheet
                ws_summary = wb.active
                ws_summary.title = "Summary"
                summary_data = [{"Metric": k, "Value": v} for k, v in report_data["summary"].items()]
                for row in summary_data:
                    ws_summary.append(row)
                
                # Detail sheets
                for key, data in report_data["details"].items():
                    ws = wb.create_sheet(title=key.capitalize())
                    if data:
                        headers = list(data[0].keys())
                        ws.append(headers)
                        for row in data:
                            ws.append([row.get(header, '') for header in headers])
                
                output = io.BytesIO()
                wb.save(output)
                return output.getvalue()
            else:
                # Export summary as Excel
                summary_data = [{"Metric": k, "Value": v} for k, v in report_data["summary"].items()]
                return self.export_excel(summary_data)
        
        elif format_type == "pdf":
            # Export as PDF
            title = f"Smart Locker System Report - {report_type.title()}"
            sections = [
                {
                    "title": "System Summary",
                    "content": [f"{k.replace('_', ' ').title()}: {v}" for k, v in report_data["summary"].items()]
                }
            ]
            
            if report_type == "comprehensive":
                for key, data in report_data["details"].items():
                    sections.append({
                        "title": f"{key.replace('_', ' ').title()}",
                        "content": data[:10]  # Limit to first 10 items for PDF
                    })
            
            return self.export_pdf(title, sections)
        
        else:
            raise ValueError(f"Unsupported format: {format_type}")

# Global export manager instance
export_manager = ExportManager()

def export_data_csv(data: List[Dict], filename: str = None) -> str:
    """Export data to CSV format"""
    return export_manager.export_csv(data, filename)

def export_data_excel(data: List[Dict], filename: str = None) -> bytes:
    """Export data to Excel format"""
    return export_manager.export_excel(data, filename)

def export_data_pdf(title: str, sections: List[Dict], filename: str = None) -> bytes:
    """Export data to PDF format"""
    return export_manager.export_pdf(title, sections, filename)

def export_system_report(db_session, format_type: str = "pdf", report_type: str = "comprehensive") -> bytes:
    """Export system report in specified format"""
    return export_manager.export_system_report(db_session, format_type, report_type) 